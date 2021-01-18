import re
import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook("License.xlsx")
ws = wb.active
# ws.append(["Name","License Key"])
rows = ws.max_row
columns = ws.max_column
for i in range(1,rows):
    cell_obj = ws.cell(row=i,column=3)
    key = str(cell_obj.value)
    res = re.search(r'[A-Z0-9]{16}',key)
    if res:
        print(ws.cell(row=i,column=1).value,"\t\t",res.group())
   

