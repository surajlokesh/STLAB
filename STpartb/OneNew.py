import openpyxl as xl;
import xlsxwriter as xs

#create new excel file
workbook = xs.Workbook('LessThan60.xlsx')
ws2 = workbook.add_worksheet()
# opening the source excel file
filename ="demo.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]
#calculate number of rows in source excel file
maxR = ws1.max_row

ws2.write(0,0,"Name")
ws2.write(0,1,'USN')
ws2.write(0,2,'Subject 1')
ws2.write(0,3,'Subject 2')
ws2.write(0,4,'Subject 3')

sr1=1
sr2=0
for i in range (2, maxR + 1):
    for j in range (3,6):
        # reading cell value from source excel file
        c = int(ws1.cell(row = i, column = j).value)
        if(c<60):
            for sr2 in range(1,6):
                ws2.write(sr1,sr2-1,ws1.cell(i,sr2).value)
            sr1=sr1+1
            break
workbook.close()